"""LibreOffice 経由で .xlsx の全数式を再計算し、エラーを検出する。

機能:
- LibreOffice Basic マクロによる数式再計算
- 再計算後のファイルを data_only で読み込みエラーをスキャン
- 7 種類の Excel エラー（#VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A）を検出
- 結果を JSON で標準出力に出力
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any

# Excel エラー値
EXCEL_ERRORS = frozenset([
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
])

MAX_LOCATIONS_PER_ERROR = 20

# LibreOffice Basic マクロ
_MACRO_CONTENT = """\
Sub RecalculateAndSave(sURL As String)
    Dim oDoc As Object
    Dim oProps(0) As New com.sun.star.beans.PropertyValue
    oProps(0).Name = "Hidden"
    oProps(0).Value = True

    oDoc = StarDesktop.loadComponentFromURL( _
        ConvertToURL(sURL), "_blank", 0, oProps())

    If IsNull(oDoc) Or IsEmpty(oDoc) Then
        MsgBox "Failed to open: " & sURL
        Exit Sub
    End If

    oDoc.calculateAll()
    oDoc.store()
    oDoc.close(True)
End Sub
"""

_MACRO_NAME = "RecalcMacro"
_MACRO_FILENAME = "RecalcMacro.xba"


def _find_soffice() -> str | None:
    """soffice のパスを検索する。"""
    path = shutil.which("soffice")
    if path:
        return path
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.isfile(mac_path):
        return mac_path
    return None


def _get_macro_dir() -> Path:
    """LibreOffice ユーザーマクロディレクトリを返す。"""
    home = Path.home()
    linux_dir = home / ".config/libreoffice/4/user/basic/Standard"
    if linux_dir.parent.parent.exists():
        return linux_dir
    mac_dir = home / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    if mac_dir.parent.parent.exists():
        return mac_dir
    return linux_dir


def _ensure_macro(soffice_path: str) -> None:
    """再計算マクロをセットアップする。"""
    macro_dir = _get_macro_dir()
    macro_file = macro_dir / _MACRO_FILENAME

    if macro_file.exists() and macro_file.read_text().strip() == _MACRO_CONTENT.strip():
        return

    if not macro_dir.exists():
        print("Initializing LibreOffice profile...", file=sys.stderr)
        subprocess.run(
            [soffice_path, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=30,
        )
        time.sleep(2)
        macro_dir.mkdir(parents=True, exist_ok=True)

    macro_file.write_text(_MACRO_CONTENT)

    script_xlb = macro_dir / "script.xlb"
    if not script_xlb.exists():
        script_xlb.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD '
            'OfficeDocument 1.0//EN" "library.dtd">\n'
            '<library:library xmlns:library='
            '"http://openoffice.org/2000/library" '
            'library:name="Standard" library:readonly="false" '
            'library:passwordprotected="false">\n'
            f'  <library:element library:name="{_MACRO_NAME}"/>\n'
            "</library:library>\n"
        )

    print(f"Macro installed: {macro_file}", file=sys.stderr)


def _run_recalc(soffice_path: str, xlsx_path: Path, timeout: int) -> None:
    """LibreOffice マクロで数式を再計算する。"""
    abs_path = str(xlsx_path.resolve())
    macro_url = f"macro:///Standard.{_MACRO_NAME}.RecalculateAndSave({abs_path})"

    cmd = [soffice_path, "--headless", "--norestore", macro_url]

    try:
        result = subprocess.run(cmd, capture_output=True, timeout=timeout, text=True)
        if result.returncode != 0 and result.stderr:
            print(f"soffice stderr: {result.stderr}", file=sys.stderr)
    except subprocess.TimeoutExpired:
        print(
            f"Error: LibreOffice timed out after {timeout}s.",
            file=sys.stderr,
        )
        sys.exit(1)


def _scan_errors(xlsx_path: Path) -> dict[str, Any]:
    """再計算後のファイルをスキャンしてエラーを検出する。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Error: openpyxl is not installed. Run: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)

    total_formulas = 0
    total_errors = 0
    error_summary: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # 数式カウント（data_only=True なので数式は値に置換されている）
                # 元ファイルから数式数をカウントするために再読み込み
                val = str(cell.value)

                if val in EXCEL_ERRORS:
                    total_errors += 1
                    if val not in error_summary:
                        error_summary[val] = {"count": 0, "locations": []}
                    error_summary[val]["count"] += 1
                    if len(error_summary[val]["locations"]) < MAX_LOCATIONS_PER_ERROR:
                        loc = f"{ws.title}!{cell.coordinate}"
                        error_summary[val]["locations"].append(loc)

    wb.close()

    # 数式カウント（data_only=False で再読み込み）
    wb2 = load_workbook(xlsx_path, data_only=False)
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
    wb2.close()

    status = "success" if total_errors == 0 else "errors_found"

    return {
        "status": status,
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": error_summary,
    }


def recalc(xlsx_path: Path, timeout: int = 60) -> dict[str, Any]:
    """数式を再計算しエラーを検出する。"""
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    suffix = xlsx_path.suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        print(f"Error: expected .xlsx or .xlsm file: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    soffice_path = _find_soffice()
    if soffice_path is None:
        print(
            "Error: LibreOffice (soffice) not found.\n"
            "Install: apt: sudo apt install libreoffice / "
            "brew: brew install --cask libreoffice",
            file=sys.stderr,
        )
        sys.exit(1)

    _ensure_macro(soffice_path)

    print("Recalculating formulas via LibreOffice...", file=sys.stderr)
    _run_recalc(soffice_path, xlsx_path, timeout)

    print("Scanning for errors...", file=sys.stderr)
    result = _scan_errors(xlsx_path)

    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="LibreOffice 経由で .xlsx の全数式を再計算し、エラーを検出する。",
        epilog="例: python recalc.py output.xlsx",
    )
    parser.add_argument("input", type=Path, help="入力 .xlsx ファイル")
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="LibreOffice タイムアウト秒数（デフォルト: 60）",
    )
    return parser


if __name__ == "__main__":
    args = build_parser().parse_args()
    result = recalc(args.input, args.timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
